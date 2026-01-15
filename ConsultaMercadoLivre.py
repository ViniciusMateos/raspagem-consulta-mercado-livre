import os
import traceback
import time
import requests
import warnings
import urllib3
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import pandas as pd
from bs4 import BeautifulSoup

warnings.filterwarnings("ignore", category=urllib3.exceptions.InsecureRequestWarning)

def ConsultaMercadoLivre(maximotentativas=5):
    tentativas = 0
    produto = ""

    while tentativas <= maximotentativas:

        while produto == "":
            produto = input("Digite um item para pesquisar no Mercado Livre: ")
            if not " " in produto:
                produto += " "
            produto = produto.replace(" ", "+")

        try:
            # Selenium headless
            chrome_options = Options()
            chrome_options.add_argument("--headless=new")
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument(
                "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36"
            )

            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()),
                options=chrome_options
            )

            produtos = []

            for pagina in range(1, 6):  # ATÃ‰ A 5Âª PÃGINA
                if pagina == 1:
                    url = f"https://lista.mercadolivre.com.br/{produto}"
                else:
                    desde = (pagina - 1) * 48 + 1
                    url = f"https://lista.mercadolivre.com.br/{produto}_Desde_{desde}"

                print(f"ðŸ“„ Coletando pÃ¡gina {pagina}: {url}")
                driver.get(url)

                time.sleep(5)  # Mercado Livre carrega via JS

                html = driver.page_source
                soup = BeautifulSoup(html, "html.parser")

                items = soup.select('li.ui-search-layout__item, li.ui-search-layout__stack')

                for item in items:
                    nome = item.find('a', class_='poly-component__title')
                    nome = nome.text.strip() if nome else ""

                    imagem = item.find('img', class_='poly-component__picture')
                    imagem = imagem['src'] if imagem else ""

                    preco = item.find('span', class_='andes-money-amount')
                    preco = preco.text.strip() if preco else ""

                    link = item.find('a', class_='poly-component__title')
                    link = link['href'] if link else ""

                    produtos.append({
                        "Nome": nome,
                        "PreÃ§o": preco,
                        "Imagem": imagem,
                        "Link": link
                    })

                time.sleep(2)  # anti-ban bÃ¡sico ðŸ˜‡

            driver.quit()

            # Tratamento de preÃ§os
            for product in produtos:
                preco = product['PreÃ§o']
                if isinstance(preco, str):
                    preco = preco.replace('R$', '').replace('.', '').replace(',', '.').strip()
                    product['PreÃ§o'] = float(preco) if preco else 0.0

            produtos_ordenados = sorted(produtos, key=lambda x: x['PreÃ§o'])

            df = pd.DataFrame(produtos_ordenados)

            excel_file = 'produtos_ordenados.xlsx'
            df.to_excel(excel_file, index=False)

            wb = load_workbook(excel_file)
            ws = wb.active

            currency_style = NamedStyle(name="currency_style", number_format='R$ #,##0.00')

            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=2).style = currency_style

            wb.save(excel_file)

            print("âœ… ARQUIVO EXCEL CRIADO COM SUCESSO")
            os.startfile(excel_file)

            return produtos_ordenados

        except Exception as e:
            print("âŒ ERRO NA CONSULTA")
            print(e)
            tentativas += 1
            time.sleep(5)

    print("ðŸš« NÃšMERO MÃXIMO DE TENTATIVAS")
    return None

consulta = ConsultaMercadoLivre()







